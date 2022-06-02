# #####################################################################################################################
# PYTHON PROGRAM FOR SPARE PARTS REVISION
# DATE: MAY 18, 2022
# VERSION: 1.0
# AUTHOR: LUIS HERNANDEZ

#######################################################################################################################
#                                               SPARE PARTS ANALYSIS
# This program compares the incoming parts lists with the required spare parts lists and outputs a document containing:
# received parts, pending parts and received parts not included in the required parts list.
#######################################################################################################################

#######################################################################################################################
# SECTION 0A: IMPORTS
#######################################################################################################################

import os
import pandas as pd
import psutil
import openpyxl
from openpyxl.styles import Alignment

#######################################################################################################################
# SECTION OB: PATHS
#######################################################################################################################

test = True

if test is True:
    path1 = r'C:\Users\Airtech\Desktop\Proyectos\Spare Parts Analysis (SPA)\test files'
    path2 = r'C:\Users\Airtech\Desktop\Proyectos\Spare Parts Analysis (SPA)\out\ANÁLISIS DE REPUESTOS.xlsx'
else:
    path1 = r'C:\Users\Airtech\Desktop\Proyectos\Spare Parts Analysis (SPA)\input'
    path2 = r'C:\Users\Airtech\Desktop\Proyectos\Spare Parts Analysis (SPA)\out\ANÁLISIS DE REPUESTOS.xlsx'

#######################################################################################################################
# SECTION 1: CREATION OF DATAFRAMES FOR INCOMING AND REQUIRED PARTS.
'''
    Objective:
        > Create two dataFrames, one for the Incoming parts and the other for the Required parts.
'''
#######################################################################################################################

# 1.1) Closes Excel program.
for proc in psutil.process_iter():
    if proc.name() == "excel.exe" or proc.name() == "EXCEL.EXE":
        proc.kill()

# 1.2) Creation of list with input files names.
fileList = os.listdir(path1)
fileNames = []
for filename in fileList:
    if filename.endswith(".xlsx") or filename.endswith(".xlsm") or filename.endswith(".xls"):
        fileNames.append(filename)

# 1.3) Creation of DataFrames for Incoming and Required parts.
frames = []
for file in fileNames:
    frames.append(pd.read_excel(path1 + '\\' + file))

#######################################################################################################################
# SECTION 2: DATAFRAMES PRE-PROCESSING.
'''
    Objective:
        > Prepare the Incoming and Required dataFrames for further analysis.
'''
#######################################################################################################################

# 2.1) P/N Correction (example: on the Incoming dataFrame the RTV104 must appear in P/N and not in Description).
incoming_df = frames[0]
for index in incoming_df.index:
    itemPN = incoming_df['PN'][index]
    incoming_df.loc[index, 'PN'] = "".join(str(itemPN).split())
    itemDesc = str(incoming_df['DESC'][index])
    desc = "".join(itemDesc.split())
    if desc.isalpha() is False:
        if (str(itemPN).isalnum() is False) or (str(itemPN).isnumeric() is False):
            incoming_df.loc[index, 'PN'] = itemDesc
            incoming_df.loc[index, 'DESC'] = ''

# 2.2) Quantity correction (example: 12 oz. should say 1).
required_df = frames[1]
for index in required_df.index:
    itemPN = required_df['PN'][index]
    required_df.loc[index, 'PN'] = "".join(str(itemPN).split())
    itemQTY = str(required_df['QTY'][index])
    QTY = "".join(itemQTY.split())
    if QTY.isnumeric() is False:
        required_df.loc[index, 'QTY'] = 1

# 2.3) Adding duplicates in Incoming list.

# 2.3.1) Indexes of duplicated items.
duplicatedIndexes, duplicated_PN = [], []
for index1 in incoming_df.index:
    indexes, partial_indexes = [], []
    partNum1 = ''.join(filter(str.isalnum, incoming_df['PN'][index1]))
    for index2 in incoming_df.index:
        if any(index2 in sublist for sublist in duplicatedIndexes) is False:
            partNum2 = ''.join(filter(str.isalnum, incoming_df['PN'][index2]))
            if partNum1 == partNum2 and index1 != index2:
                indexes.append(index2)
    if len(indexes) != 0:
        partial_indexes = [index1] + indexes
        duplicatedIndexes.append(partial_indexes)

# 2.3.2) Total quantity of duplicated items are added in a new row. Original row is deleted.
newRows = []
lastNum = frames[0]['#'][frames[0].index[-1]]
i = 0
for index_group in duplicatedIndexes:
    totalQTY = 0
    for index in index_group:
        totalQTY = totalQTY + frames[0]['QTY'][index]
    duplicatedDESC = frames[0]['DESC'][index_group[0]]
    duplicatedPN = frames[0]['PN'][index_group[0]]
    i = i + 1
    newRows.append(pd.DataFrame({'#': [lastNum+i], 'DESC': [duplicatedDESC], 'PN': [duplicatedPN], 'QTY': [totalQTY]}))

# 2.3.3) Original rows of duplicated items are deleted.
inFrame = frames[0].copy(deep=True)
total_duplicated_indexes = []
for index_group in duplicatedIndexes:
    for index in index_group:
        total_duplicated_indexes.append(index)
inFrame.drop(total_duplicated_indexes, inplace=True)

# 2.3.4) newRows are appended to the incoming dataFrame.
for row in newRows:
    inFrame = pd.concat([inFrame, row], ignore_index=True, axis=0)

########################################################################################################################
# SECTION 3: INCOMING AND REQUIRED PARTS ANALYSIS.
'''
    Objective:
        > Create xlsx files for: received parts, pending parts and parts that weren't in the original parts list.
'''
########################################################################################################################

# 3.1) DataFrames comparison considering P/N and Received/Required quantities.
receivedDict = {}
missing = frames[1].copy(deep=True)
extra = inFrame.copy(deep=True)
req_idx, rec_idx = [], []
key = 0
for ind1 in frames[1].index:
    reqItem = ''.join(filter(str.isalnum, frames[1]['PN'][ind1]))
    for ind2 in inFrame.index:
        inItem = ''.join(filter(str.isalnum, inFrame['PN'][ind2]))
        if reqItem == inItem:
            key = key + 1
            PN = frames[1]['PN'][ind1]
            reqQTY = frames[1]['QTY'][ind1]
            inQTY = inFrame['QTY'][ind2]
            difQTY = inQTY - reqQTY
            req_idx.append(ind1)
            rec_idx.append(ind2)
            values = [PN, reqQTY, inQTY, difQTY]
            receivedDict[key] = values

missing.drop(index=req_idx, inplace=True)
extra.drop(index=rec_idx, inplace=True)

missing.reset_index(drop=True, inplace=True)
extra.reset_index(drop=True, inplace=True)
received = pd.DataFrame.from_dict(receivedDict, orient='index', columns=['P/N', 'QTY. PEDIDO', 'QTY. RECIBIDO', 'DIF.'])

# 3.2) Setting column width and saving dataFrames as .xlsx files.
writer = pd.ExcelWriter(path2, engine='xlsxwriter')

received.to_excel(writer, sheet_name='Recibido', index=False)
missing.to_excel(writer, sheet_name='Faltante', index=False)
extra.to_excel(writer, sheet_name='No pedido', index=False)

worksheets = ['Recibido', 'Faltante', 'No pedido']
df_list = [received, missing, extra]
i = 0
for df in df_list:
    sheetname = worksheets[i]
    for column in df:
        column_width = max(df[column].astype(str).map(len).max(), len(column))
        col_idx = df.columns.get_loc(column)
        writer.sheets[sheetname].set_column(col_idx, col_idx, column_width * 1.1)
    i += 1

########################################################################################################################
# SECTION 4: EXCEL FILES FORMATTING.
'''
    Objective:
        > Add format to .xlsx files.
'''
########################################################################################################################
finalFrames = [received, missing, extra]

# 4.1)  Add color to header of table.
workbook = writer.book
header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'fg_color': '#59BFFF', 'border': 1})
for (sheet, frame) in zip(worksheets, finalFrames):
    worksheet = writer.sheets[sheet]
    for col_num, value in enumerate(frame.columns.values):
        worksheet.write(0, col_num, value, header_format)

# 4.2) Add borders to table and set column width.
for (sheet, frame) in zip(worksheets, finalFrames):
    worksheet = writer.sheets[sheet]
    row_idx, col_idx = frame.shape
    for r in range(row_idx):
        for c in range(col_idx):
            values = frame.values[r, c]
            worksheet.write(r+1, c, frame.values[r, c], workbook.add_format({'border': 1}))
    worksheet.set_column(0, 4, 15)

writer.close()

# 4.3) Center align cells of the .xlsx file.
wb = openpyxl.load_workbook(path2)
center_aligned_text = Alignment(horizontal="center")
for sheet in wb.worksheets:
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            cell_obj = sheet.cell(row=i, column=j)
            coordinate = cell_obj.coordinate
            sheet[coordinate].alignment = center_aligned_text

wb.save(path2)
